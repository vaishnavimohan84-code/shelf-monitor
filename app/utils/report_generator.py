"""
Report generation for the AI Product Shelf Monitoring System (Step 12).

Produces PDF, Excel, and CSV exports summarizing detection scans and alerts
for a given date range, saved into app/static/reports/.
"""
import csv
import os
import uuid
from datetime import datetime

from app.models.detection import DetectionLog, Alert


def _filtered_logs(start_date=None, end_date=None, shelf_id=None):
    query = DetectionLog.query
    if start_date:
        query = query.filter(DetectionLog.created_at >= start_date)
    if end_date:
        query = query.filter(DetectionLog.created_at <= end_date)
    if shelf_id:
        query = query.filter(DetectionLog.shelf_id == shelf_id)
    return query.order_by(DetectionLog.created_at.desc()).all()


def _filtered_alerts(start_date=None, end_date=None, shelf_id=None):
    query = Alert.query
    if start_date:
        query = query.filter(Alert.created_at >= start_date)
    if end_date:
        query = query.filter(Alert.created_at <= end_date)
    if shelf_id:
        query = query.filter(Alert.shelf_id == shelf_id)
    return query.order_by(Alert.created_at.desc()).all()


def generate_csv_report(reports_folder, start_date=None, end_date=None, shelf_id=None):
    logs = _filtered_logs(start_date, end_date, shelf_id)

    filename = f"scan_report_{uuid.uuid4().hex[:8]}.csv"
    filepath = os.path.join(reports_folder, filename)

    with open(filepath, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Scan ID", "Shelf", "Capture Type", "Products Detected",
            "Empty Shelf %", "Is Empty", "Scanned At",
        ])
        for log in logs:
            writer.writerow([
                log.id,
                log.shelf.name,
                log.capture_type,
                log.total_products_detected,
                log.empty_shelf_percentage,
                "Yes" if log.is_empty else "No",
                log.created_at.strftime("%Y-%m-%d %H:%M"),
            ])

    return filename


def generate_excel_report(reports_folder, start_date=None, end_date=None, shelf_id=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    logs = _filtered_logs(start_date, end_date, shelf_id)
    alerts = _filtered_alerts(start_date, end_date, shelf_id)

    wb = Workbook()

    # --- Scans sheet ---
    ws = wb.active
    ws.title = "Scans"
    headers = ["Scan ID", "Shelf", "Capture Type", "Products Detected", "Empty Shelf %", "Is Empty", "Scanned At"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0D6EFD")

    for log in logs:
        ws.append([
            log.id, log.shelf.name, log.capture_type, log.total_products_detected,
            log.empty_shelf_percentage, "Yes" if log.is_empty else "No",
            log.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
        ws.column_dimensions[column_cells[0].column_letter].width = max(12, length + 2)

    # --- Alerts sheet ---
    ws2 = wb.create_sheet("Alerts")
    ws2.append(["Alert ID", "Shelf", "Type", "Message", "Resolved", "Created At"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="DC3545")

    for alert in alerts:
        ws2.append([
            alert.id, alert.shelf.name, alert.alert_type, alert.message,
            "Yes" if alert.is_resolved else "No",
            alert.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    for column_cells in ws2.columns:
        length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
        ws2.column_dimensions[column_cells[0].column_letter].width = max(12, length + 2)

    filename = f"shelf_report_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(reports_folder, filename)
    wb.save(filepath)
    return filename


def generate_pdf_report(reports_folder, start_date=None, end_date=None, shelf_id=None):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    logs = _filtered_logs(start_date, end_date, shelf_id)
    alerts = _filtered_alerts(start_date, end_date, shelf_id)

    filename = f"shelf_report_{uuid.uuid4().hex[:8]}.pdf"
    filepath = os.path.join(reports_folder, filename)

    doc = SimpleDocTemplate(filepath, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("AI Product Shelf Monitoring — Report", styles["Title"]))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    elements.append(Spacer(1, 0.5 * cm))

    elements.append(Paragraph(f"Scan Summary ({len(logs)} scans)", styles["Heading2"]))
    scan_data = [["ID", "Shelf", "Type", "Detected", "Empty %", "Empty?", "Scanned At"]]
    for log in logs[:200]:  # cap rows to keep PDF manageable
        scan_data.append([
            str(log.id), log.shelf.name, log.capture_type, str(log.total_products_detected),
            f"{log.empty_shelf_percentage}%", "Yes" if log.is_empty else "No",
            log.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    scan_table = Table(scan_data, repeatRows=1)
    scan_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D6EFD")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F9")]),
    ]))
    elements.append(scan_table)
    elements.append(Spacer(1, 0.8 * cm))

    elements.append(Paragraph(f"Alerts ({len(alerts)} total)", styles["Heading2"]))
    alert_data = [["ID", "Shelf", "Type", "Message", "Created At"]]
    for alert in alerts[:200]:
        alert_data.append([
            str(alert.id), alert.shelf.name, alert.alert_type, alert.message,
            alert.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    alert_table = Table(alert_data, repeatRows=1, colWidths=[1.5 * cm, 3 * cm, 3 * cm, 7 * cm, 3.5 * cm])
    alert_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DC3545")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F9")]),
    ]))
    elements.append(alert_table)

    doc.build(elements)
    return filename
